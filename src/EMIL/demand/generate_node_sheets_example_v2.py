from __future__ import annotations

from pathlib import Path
import re
import sys

import openpyxl
import pandas as pd

# ---------------------------------------------------------------------------

NODE_LIST: list[str] = [
    "ALh2", "ATh2", "ATh2Z1", "BAh2", "BEh2", "BEH2Mo", "BEh2Z1", "BGh2", "BGh2Z1", "CHh2", "CYh2", "CZh2", "CZh2Z1",
    "DEh2", "DEh2ba", "DEh2bp", "DEh2Z1a", "DEh2Z1b", "DKh2", "DKh2Z1", "EEh2", "EEh2Z1", "ESh2", "ESh2Z1a", "ESh2Z1b",
    "FIh2", "FIh2Al", "FIh2N", "FIh2S", "FIh2SZ1", "FRh2", "FRh2N", "FRh2S", "FRh2SW", "FRh2Va", "FRh2Z1", "GRh2",
    "GRh2Z1", "HRh2", "HRh2Z1", "HUh2", "HUh2Z1", "IB-ITh2", "IB-SKh2C", "IB-SKh2E", "IB-SKh2W", "IEh2", "IEh2Z1",
    "ITh2", "ITh2Z1", "LTh2", "LTh2Z1", "LUh2", "LUh2Z1", "LVh2", "LVh2Z1", "MDh2", "MKh2", "MTh2", "NLh2", "NLh2Z1",
    "NOhe", "PLh2N", "PLh2nbc", "PLh2NZ1a", "PLh2NZ1b", "PLh2S", "PLh2SZ1a", "PLh2SZ1b", "PTh2", "PTh2Z1", "ROh2",
    "ROh2Z1", "RSh2", "SEh2", "SEh2Z1", "SIh2", "SIh2Z1", "SKh2E", "SKh2EZ1", "SKh2W", "SKh2WZ1", "UKh2", "UKh2/INT",
    "UKh2Z1", "Y-NOh2",
]

ELECTRICITY_NODE_LIST: list[str] = [
                                    "AL00", "BA00", "BE00", "BG00", "CH00", "CY00", "CZ00", "DE00", "DKE1", "DKW1", "EE00", "ES00", "FI00",
                                    "FR00", "GR00", "GR03", "HR00", "HU00", "IE00", "ITCA", "ITCN", "ITCS", "ITN1", "ITS1", "ITSA", "ITSI", "LT00",
                                    "LUG1", "LV00", "MD00", "ME00", "MK00", "MT00", "NL00", "NOM1", "NON1", "NOS1", "NOS2", "NOS3", "PL00", "PT00",
                                    "RO00", "RS00", "SE01", "SE02", "SE03", "SE04", "SI00", "SK00", "TR00", "UA00", "UKNI"
                                ]

ELECTRICITY_H2_NODE_LIST: list[str] = [
                                    "AL00", "BA00", "BE00", "BG00", "CH00", "CY00", "CZ00", "DE00", "DKE1", "DKW1", "EE00", "ES00", "FIh2N", "FIh2S",
                                    "FRh2S", "FRh2SW","GR00", "GR03", "HR00", "HU00", "IE00", "ITCA", "ITCN", "ITCS", "ITN1", "ITS1", "ITSA", "ITSI", "LT00",
                                    "LUG1", "LV00", "MD00", "ME00", "MK00", "MT00", "NL00", "NOM1", "NON1", "NOS1", "NOS2", "NOS3", "PLh2N", "PLh2S", "PT00",
                                    "RO00", "RS00", "SE01", "SE02", "SE03", "SE04", "SI00", "SKh2E", "SKh2W", "TR00", "UA00", "UKNI"
                                ]

NODE_LIST_Z1 = [n for n in NODE_LIST if "Z1" in n]
NODE_LIST_Z2 = [n for n in NODE_LIST if "Z1" not in n]

# Folder containing the template workbooks
ROOT_FOLDER = Path(r"C:\Users\ENTSOE\Tera-joule\Terajoule - Terajoule\Projects\ENTSOG\Scenarios\Demand Profiling\Demand 2026 Examples")
YEARS = [2030, 2035, 2040, 2050]
SCENARIO = "NT"

file_node_dict = {"CH4 HEAT DEMAND": ELECTRICITY_NODE_LIST,
                  "H2 HEAT DEMAND": ELECTRICITY_H2_NODE_LIST,
                  "H2_ZONE_1": NODE_LIST_Z1,
                  "H2_ZONE_2": NODE_LIST_Z2,
                  "SYNTHETIC FUELS METHANE": NODE_LIST_Z2, 
                  "SYNTHETIC FUELS LIQUIDS": NODE_LIST_Z2,                   
                  }
                  

# Create a template dictionary for each year using openpyxl
TEMPLATE_FILES = {
    2030: r"C:\Users\ENTSOE\Tera-joule\Terajoule - Terajoule\Projects\ENTSOG\Scenarios\Demand Profiling\Demand 2026 Examples\samples\ELECTRICITY_MARKET 2030.xlsx",
    2035: r"C:\Users\ENTSOE\Tera-joule\Terajoule - Terajoule\Projects\ENTSOG\Scenarios\Demand Profiling\Demand 2026 Examples\samples\ELECTRICITY_MARKET 2035.xlsx",
    2040: r"C:\Users\ENTSOE\Tera-joule\Terajoule - Terajoule\Projects\ENTSOG\Scenarios\Demand Profiling\Demand 2026 Examples\samples\ELECTRICITY_MARKET 2040.xlsx",
    2050: r"C:\Users\ENTSOE\Tera-joule\Terajoule - Terajoule\Projects\ENTSOG\Scenarios\Demand Profiling\Demand 2026 Examples\samples\ELECTRICITY_MARKET 2050.xlsx",
}

template_wb_dict = {}
for year, file_path in TEMPLATE_FILES.items():
    wb = openpyxl.load_workbook(file_path)
    template_wb_dict[year] = wb


# ---------------------------------------------------------------------------

def safe_sheet_name(name: str) -> str:
    """Return *name* as a valid Excel sheet title (≤31 chars, no invalid symbols)."""
    invalid = r'[:\\/?*\[\]]'
    cleaned = re.sub(invalid, '-', name)
    return cleaned[:31]  # truncate to 31 chars


def update_metadata(ws, node_name: str, year: int) -> None:
    """Update metadata in a worksheet."""
    ws["B3"] = node_name  # Node name
    ws["B5"] = year       # Target Year
    ws["B6"] = SCENARIO   # Scenario


def add_node_sheets(wb: openpyxl.Workbook, file_name: str, year: int) -> None:
    """Duplicate 'AT00' template for every node and rename the copies accordingly."""
    template_ws = wb["AT00"]

    NODE_LIST = file_node_dict.get(file_name, [])
    for node in NODE_LIST:
        safe_name = safe_sheet_name(node)
        new_sheet = wb.copy_worksheet(template_ws)
        new_sheet.title = safe_name
        # Update metadata for this specific node
        update_metadata(new_sheet, node, year)


def remove_unneeded_sheets(wb: openpyxl.Workbook) -> None:
    """Delete 'AT00' and 'DE00' sheets if present."""
    for name in ("AT00", "DE00"):
        if name in wb.sheetnames:
            ws = wb[name]
            wb.remove(ws)


def process_workbook(path: Path, year: int) -> None:
    file_name = path.stem
    print(f"Processing {path.name} …")
    
    # Start with a fresh copy of the template workbook for this year
    template_path = TEMPLATE_FILES[year]
    wb = openpyxl.load_workbook(template_path)
    
    # Create new output directory with suffix "_updated"
    output_dir = path.parent / f"{path.parent.name}_updated"
    output_dir.mkdir(exist_ok=True)
    new_path = output_dir / path.name
    try:
        add_node_sheets(wb, file_name, year)
        remove_unneeded_sheets(wb)
        wb.save(new_path)
    except Exception as exc:
        print(f"✗ {path.name}: {exc}")
    else:
        print(f"✓ {new_path} updated.")


def main() -> None:
    for year in YEARS:
        FOLDER = ROOT_FOLDER / str(year)
        if not FOLDER.exists():
            sys.exit(f"Folder not found: {FOLDER}")

        workbooks = list(FOLDER.glob("*.xlsx"))  # adapt pattern if needed
        if not workbooks:
            sys.exit("No .xlsx files found.")

        for wb_path in workbooks:
            process_workbook(wb_path, year)


if __name__ == "__main__":
    main()
